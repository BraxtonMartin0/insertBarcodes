import sys
from openpyxl import load_workbook


def insertBarcodes(file, gridsFile):
    wb = load_workbook(file)
    data = wb['Sheet1']

    allClasses = []

    count = data.max_row
    x = 2
    while x < count - 1:
        info_cell = "A" + str(x)
        allInfo = data[info_cell].value
        info = allInfo.split("-")

        activityNumber = info[0]
        if len(info) > 2:
            swimClassName = "PVL"
            level = "PVL"
        else:
            swimClassNameInfo = info[1].split("–")
            swimClassName = swimClassNameInfo[0]
            level = swimClassNameInfo[1]

        if "|" in level:
            temp = level.split("|")
            level = temp[1]

        if "Private Lesson" in level and not "Semi" in level:
            level = "PVL"
        elif "Private Lesson" in level and "Semi" in level:
            level = "Semi-PVL"

        if "Little" in level:
            tempLevel = level.split("Little")
            level = tempLevel[1].replace(" ", "")

        if "Canadian Tire Jumpstart I Love to Swim" in level:
            level = "I Love to Swim"

        if "women" in level.lower():
            tempLevel = level.split(" ")
            level = tempLevel[0] + " women"

        if "low ratio" in level:
            lowRatio = True
            levelTemp1 = level.split("(")
            level = levelTemp1[0]
        else:
            lowRatio = False

        # print(level)
        day_cell = "D" + str(x)
        day = data[day_cell].value

        time_cell = "E" + str(x)
        timeTemp = data[time_cell].value.split(":")

        if "PM" in timeTemp[1]:
            PM = True
        else:
            PM = False

        hour = timeTemp[0]
        minute = timeTemp[1][:2]

        newClass = SwimClass(activityNumber, swimClassName, level, day, hour, minute, PM, lowRatio)

        allClasses.append(newClass)

        x += 1

    print("There are " + str(len(allClasses)) + " classes this session")

    wb.save(file)

    missingClasses = []

    for swimmingClass in allClasses:
        gridsWb = load_workbook(gridsFile)

        if swimmingClass.PM and not int(swimmingClass.hour) == 12:
            classTime = str(int(swimmingClass.hour) + 12) + ":" + str(swimmingClass.minute) + ":" + "00"
        else:
            classTime = str(swimmingClass.hour) + ":" + str(swimmingClass.minute)

        if swimmingClass.day == "Sa":
            grid = gridsWb["Saturday AM"]
            daytime = False
            if not (gridSearch(grid, classTime, swimmingClass.level, swimmingClass.activityNumber,
                               swimmingClass.lowRatio,
                               daytime)):
                missingClasses.append(swimmingClass)
            gridsWb.save(gridsFile)
        elif swimmingClass.day == "Su":
            grid = gridsWb["Sunday AM"]
            daytime = False
            if not (gridSearch(grid, classTime, swimmingClass.level, swimmingClass.activityNumber,
                               swimmingClass.lowRatio,
                               daytime)):
                missingClasses.append(swimmingClass)
            gridsWb.save(gridsFile)
        elif swimmingClass.day == "M":
            if int(swimmingClass.hour) > 1 and swimmingClass.PM and not (int(swimmingClass.hour) == 12):
                grid = gridsWb["Monday PM"]
                daytime = False
                if not (gridSearch(grid, classTime, swimmingClass.level, swimmingClass.activityNumber,
                                   swimmingClass.lowRatio,
                                   daytime)):
                    missingClasses.append(swimmingClass)
                gridsWb.save(gridsFile)
            else:
                grid = gridsWb["Daytime"]
                daytime = True

        elif swimmingClass.day == "Tu":
            if int(swimmingClass.hour) > 3 and swimmingClass.PM and not (int(swimmingClass.hour) == 12):
                grid = gridsWb["Tuesday PM"]
                daytime = False
                if not (gridSearch(grid, classTime, swimmingClass.level, swimmingClass.activityNumber,
                                   swimmingClass.lowRatio,
                                   daytime)):
                    missingClasses.append(swimmingClass)
                gridsWb.save(gridsFile)
            else:
                grid = gridsWb["Daytime"]
                daytime = True

        elif swimmingClass.day == "W":
            if int(swimmingClass.hour) > 3 and swimmingClass.PM and not (int(swimmingClass.hour) == 12):
                grid = gridsWb["Wednesday PM"]
                daytime = False
                if not (gridSearch(grid, classTime, swimmingClass.level, swimmingClass.activityNumber,
                                   swimmingClass.lowRatio,
                                   daytime)):
                    missingClasses.append(swimmingClass)
                gridsWb.save(gridsFile)
            else:
                grid = gridsWb["Daytime"]
                daytime = True

        elif swimmingClass.day == "Th":
            if int(swimmingClass.hour) > 3 and swimmingClass.PM and not (int(swimmingClass.hour) == 12):
                grid = gridsWb["Thursday PM"]
                daytime = False
                if not (gridSearch(grid, classTime, swimmingClass.level, swimmingClass.activityNumber,
                                   swimmingClass.lowRatio,
                                   daytime)):
                    missingClasses.append(swimmingClass)
                gridsWb.save(gridsFile)
            else:
                grid = gridsWb["Daytime"]
                daytime = True

        elif swimmingClass.day == "F":
            if int(swimmingClass.hour) > 3 and swimmingClass.PM and not (int(swimmingClass.hour) == 12):
                grid = gridsWb["Friday PM"]
                daytime = False
                if not (gridSearch(grid, classTime, swimmingClass.level, swimmingClass.activityNumber,
                                   swimmingClass.lowRatio,
                                   daytime)):
                    missingClasses.append(swimmingClass)
                gridsWb.save(gridsFile)
            else:
                grid = gridsWb["Daytime"]
                daytime = True

        else:
            print("Invalid day")

    print("")
    print("")
    print("All missed classes:")
    for swimClass in missingClasses:
        print(
            swimClass.swimClassName + " - " + swimClass.level + " - " + swimClass.day + " - " + swimClass.activityNumber + " - " + swimClass.hour + ":" + swimClass.minute)


def gridSearch(sheet, time, level, activityNumber, lowRatio, daytime):
    letters = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W",
               "X", "Y", "Z"]

    timeFound = False
    classNameFound = False
    x = 0
    classNameRow = 7
    activityNumberRow = 8
    end = False
    if lowRatio:
        print("Looking for " + level + " LR - " + activityNumber + " " + time + " " + str(sheet))
    else:
        print("Looking for " + level + " - " + activityNumber + " " + time + " " + str(sheet))
    while not timeFound or end:
        gridTimeCell = sheet[letters[x] + "6"]
        if time in str(gridTimeCell.value):
            # print("found time")
            while True:
                classNameValue = sheet[letters[x] + str(classNameRow)]
                if classNameValue.value is not None:
                    if not "Lifeguarding" in str(classNameValue.value) or not "LG" in str(classNameValue.value):
                        activityNumberValue = sheet[letters[x] + str(activityNumberRow)]
                        # print(level + " - " + str(classNameValue.value))
                        if level.replace("(", "").replace(")", "").replace(" ", "").lower() in str(
                                classNameValue.value).replace(" ", "").lower() and activityNumberValue.value is None:

                            if lowRatio:
                                if "LR" in str(classNameValue.value):
                                    sheet[letters[x] + str(activityNumberRow)].value = activityNumber
                                    print("Found class - Entering barcode for: " + level + " - " + activityNumber)
                                    return True
                            else:
                                sheet[letters[x] + str(activityNumberRow)].value = activityNumber
                                print("Found class - Entering barcode for: " + level + " - " + activityNumber)
                                return True
                        elif "women" in level.lower():
                            tempLevel = level.replace("(", "").replace(")", " ").split(" ")
                            tempLevel[0].lower()
                            tempLevel[1].lower()
                            if tempLevel[0] in str(classNameValue.value).replace(" ", "").lower() and tempLevel[
                                1] in str(classNameValue.value).replace(" ", "").lower():
                                sheet[letters[x] + str(activityNumberRow)].value = activityNumber
                                print("Found class - Entering barcode for: " + level + " - " + activityNumber)
                                return True
                classNameRow += 2
                activityNumberRow += 2
                if classNameRow > 200 or activityNumberRow > 200:
                    print("Class not found")
                    return False

        if letters[x] == "Z" and not timeFound:
            print("Not Found time")
            end = True
        x += 1


class SwimClass:
    def __init__(self, activityNumber, swimClassName, level, day, hour, minute, PM, lowRatio):
        self.activityNumber = activityNumber
        self.swimClassName = swimClassName
        self.level = level
        self.day = day
        self.hour = hour
        self.minute = minute
        self.lowRatio = lowRatio
        self.PM = PM


if __name__ == '__main__':
    insertBarcodes(sys.argv[1], sys.argv[2])